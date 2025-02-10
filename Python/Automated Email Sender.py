import smtplib
import openpyxl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime
import os
import json

script_dir = os.path.dirname(__file__)
CONFIG_PATH = os.path.abspath(os.path.join(script_dir, '..', 'configFiles', 'automatedEmailSenderConfig.json'))

with open(CONFIG_PATH, 'r') as config_file:
    config = json.load(config_file)

onedrive_path = os.path.expanduser(config['onedrive_path'])
EXCEL_FILE = os.path.join(onedrive_path, 'Brands Email List for MusaddiqLakhana.xlsx')
SOURCE_SHEET = "EMAILS TO SEND"
DEST_SHEET = "SENT EMAILS"
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587 
EMAIL_ADDRESS = os.getenv("MusaddiqLakhanaGmail")
EMAIL_PASSWORD = os.getenv("MusaddiqLakhanaGmailPassword")

HEADER_ROW = ["Category", "Recipient Name", "Brand Name", "Brand Email", "Admire your", "Status"]
DEST_HEADER_ROW = HEADER_ROW + ["Sent Date and Time"]

def initialize_workbook():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = SOURCE_SHEET
        ws1.append(HEADER_ROW)
        ws2 = wb.create_sheet(DEST_SHEET)
        ws2.append(DEST_HEADER_ROW)
    else:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        for sheet_name, header in [(SOURCE_SHEET, HEADER_ROW), (DEST_SHEET, DEST_HEADER_ROW)]:
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                ws.append(header)
    
    wb.save(EXCEL_FILE)
    return wb

def send_email(to_email, subject, body):
    try:
        msg = MIMEMultipart()
        msg["From"] = EMAIL_ADDRESS
        msg["To"] = to_email
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain"))

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            server.sendmail(EMAIL_ADDRESS, to_email, msg.as_string())
        
        print(f"Email sent to {to_email}")
        return True
    except Exception as e:
        print(f"Failed to send email to {to_email}: {e}")
        return False

def move_to_sent_table(row, wb):
    dest_ws = wb[DEST_SHEET]
    dest_ws.append(list(row[:-1]) + ['SENT', datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
    wb.save(EXCEL_FILE)
    print(f"Row moved to '{DEST_SHEET}'.")

def process_excel():
    try:
        wb = initialize_workbook()
        source_ws = wb[SOURCE_SHEET]
        rows_to_delete = []

        for index, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[5] and row[5].strip().upper() == 'NEW':
                category, recipient_name, brand_name, brand_email, admire_your, _ = row
                
                if category not in config.get("email_templates", {}):
                    print(f"No template found for category: {category}")
                    continue
                
                template = config["email_templates"][category]
                subject = template["email_subject"].format(Brand_Name=brand_name)
                body = template["email_body"].format(
                    Recipient_Name=recipient_name,
                    Brand_Name=brand_name,
                    Admire_your=admire_your
                )
                
                if send_email(brand_email, subject, body):
                    move_to_sent_table(row, wb)
                    rows_to_delete.append(index)
        
        for row_index in sorted(rows_to_delete, reverse=True):
            source_ws.delete_rows(row_index)
        
        wb.save(EXCEL_FILE)

    except Exception as e:
        exception_type = type(e).__name__
        exception_detail = str(e)
        print(f"Error occurred: {exception_type} - {exception_detail}")
        send_email(os.getenv("GMAIL_ADDRESS"), "Error in Automated Email Sender Script", 
                f"An error occurred in the script:\n\nException Type: {exception_type}\nException Detail: {exception_detail}")

if __name__ == "__main__":
    process_excel()
