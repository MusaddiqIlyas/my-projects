import os
import requests
import openpyxl
from ics import Calendar
from datetime import datetime, timedelta
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import traceback
import json

# Get the directory where the script is located (using __file__)
script_dir = os.path.dirname(__file__)

# Navigate back one level and into the 'configFiles' folder
CONFIG_PATH = os.path.join(script_dir, '..', 'configFiles', 'housematesConfig.json')

# Normalize the path to make it absolute
config_file_path = os.path.abspath(CONFIG_PATH)

with open(CONFIG_PATH, 'r') as config_file:
    config = json.load(config_file)

housemates = config['housemates']
onedrive_path = os.path.expanduser(config['onedrive_path'])
calendar_url = config['calendar_url']
bin_collection_link = config['bin_collection_link']

file_path = os.path.join(onedrive_path, 'Bin Assignments Log.xlsx')

# Fetch and parse the calendar
response = requests.get(calendar_url)
calendar = Calendar(response.text)

# Get current date and calculate the date 4 weeks from now
current_date = datetime.now().date()
four_weeks_later = current_date + timedelta(weeks=4)

# Extract event dates and collect their descriptions (Collection Names)
collection_dates = {}
for event in calendar.events:
    event_date = event.begin.date()  # Only the date part
    if current_date <= event_date <= four_weeks_later:
        description = event.name
        if event_date not in collection_dates:
            collection_dates[event_date] = []
        collection_dates[event_date].append(description)

# Helper function to convert Excel date to Python date
def convert_excel_date(excel_date):
    if isinstance(excel_date, datetime):
        return excel_date.date()
    elif isinstance(excel_date, str):
        # Try converting the string to a date
        try:
            return datetime.strptime(excel_date, "%d/%m/%Y").date()
        except ValueError:
            # In case the string is in another format
            return datetime.strptime(excel_date, "%Y-%m-%d").date()  # Adjust the format accordingly
    return excel_date  # Return as is if it cannot be converted

# Helper function to format date as dd/mm/yyyy
def format_date(date_obj):
    if isinstance(date_obj, datetime):
        return date_obj.strftime('%d/%m/%Y')
    elif isinstance(date_obj, str):
        # If it's already a string, return as is
        return date_obj
    return date_obj.strftime('%d/%m/%Y')  # Handle datetime.date objects too

# Send email function for exception notifications and regular reminders
def send_email(to_email, subject, body):
    EMAIL_ADDRESS = os.getenv("GMAIL_ADDRESS")
    EMAIL_PASSWORD = os.getenv("GMAIL_PASSWORD")
    
    msg = MIMEMultipart()
    msg['From'] = EMAIL_ADDRESS
    msg['To'] = to_email
    msg['Subject'] = subject
    
    msg.attach(MIMEText(body, 'plain'))
    
    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            server.sendmail(EMAIL_ADDRESS, to_email, msg.as_string())
            print(f"Email sent to {to_email}")
    except Exception as e:
        print(f"Error sending email to {to_email}: {e}")

# Main script logic wrapped in a try-except block
try:
    # Check if the file exists, create if not
    if not os.path.exists(file_path):
        # Create a new workbook
        wb = openpyxl.Workbook()

        # Create 'Assignments' sheet and add headers
        ws1 = wb.active
        ws1.title = 'Bin Assignments'
        ws1.append([
            "Collection Date", "Assigned Person", "Email", "Collection Names", "Reminder 1 Week", "Reminder 1 Day"
        ])

        # Create 'Housemates' sheet and add headers
        ws2 = wb.create_sheet('Housemates')
        ws2.append(["Name", "Email", "On Holiday"])
        for housemate in housemates:
            # Assume "On Holiday" status is False by default
            ws2.append([housemate["name"], housemate["email"], "No"])


        # Create 'Count' sheet and add headers
        ws3 = wb.create_sheet('Bin Count')
        ws3.append(["Housemate", "Count"])
        for housemate in housemates:
            ws3.append([housemate["name"], 0])  # Initialize count as 0

        # Create 'Status' sheet to store the last assigned person index
        ws4 = wb.create_sheet('Bin Status')
        ws4.append(["Last Assigned Index"])
        ws4.append([0])

        ws5 = wb.create_sheet('Item Requests')
        ws5.append(["Requested By", "Item Requested", "Date & Time of Request", "Name", "Email", "Email Sent?"])

        ws6 = wb.create_sheet('Dishwashing Liquid Status')
        ws6.append(["Last Assigned Index"])
        ws6.append([0]) 
        
        ws7 = wb.create_sheet('Black Bin Bag Status')
        ws7.append(["Last Assigned Index"])
        ws7.append([0])

        ws8 = wb.create_sheet('Food Waste Bin Bag Status')
        ws8.append(["Last Assigned Index"])
        ws8.append([0])

        # Save the workbook
        wb.save(file_path)
    else:
        wb = openpyxl.load_workbook(file_path)

    ws1 = wb['Bin Assignments']
    ws2 = wb['Housemates']
    ws3 = wb['Bin Count']
    ws4 = wb['Bin Status']
    ws5 = wb['Item Requests']
    ws6 = wb['Dishwashing Liquid Status']
    ws7 = wb['Black Bin Bag Status']


    # Get the last assigned person index from the Status sheet
    bin_last_assigned_idx = ws4.cell(row=2, column=1).value

    # Get existing data in the Assignments sheet and collect the collection dates
    existing_dates = set()
    for row in ws1.iter_rows(min_row=2, max_col=1, values_only=True):
        # Convert the Excel date to Python date and format it as dd/mm/yyyy
        excel_date = row[0]
        existing_dates.add(format_date(convert_excel_date(excel_date)))

    # Sort the collection dates in ascending order
    sorted_dates = sorted(collection_dates.items())

    # Add new events to the 'Assignments' sheet if they do not already exist
    for event_date, collection_names in sorted_dates:
        formatted_date = format_date(event_date)

        # If the date is not in the existing dates, add it
        if formatted_date not in existing_dates:
            collection_names_str = ', '.join(collection_names)

            # Find the next housemate who is not on holiday
            assigned_person = None
            while True:
                potential_person = housemates[bin_last_assigned_idx]
                # Check if the person is on holiday in the Housemates sheet
                for row in ws2.iter_rows(min_row=2, values_only=True):
                    if row[0] == potential_person["name"] and row[2] == "No":
                        assigned_person = potential_person
                        break

                if assigned_person:
                    break
                else:
                    # Skip to the next housemate if current one is on holiday
                    bin_last_assigned_idx = (bin_last_assigned_idx + 1) % len(housemates)

            if not assigned_person:
                print("Error: No available housemates for assignment.")
                continue

            ws1.append([
                formatted_date,  # Insert the date in dd/mm/yyyy format
                assigned_person["name"],
                assigned_person["email"],
                collection_names_str,  # All events for that day in one cell
                'Not Sent',  # Reminder 1 Week
                'Not Sent'   # Reminder 1 Day
            ])

            # Move to the next housemate
            bin_last_assigned_idx = (bin_last_assigned_idx + 1) % len(housemates)

            # Mark this date as processed
            existing_dates.add(formatted_date)

    # Update the 'Status' sheet with the new index of the last assigned person
    ws4.cell(row=2, column=1, value=bin_last_assigned_idx)

    # Update the 'Count' sheet with the number of assignments for each housemate
    ws3 = wb['Bin Count']
    housemate_count = {housemate["name"]: 0 for housemate in housemates}

    # Count how many times each housemate has been assigned in the 'Assignments' sheet
    for row in ws1.iter_rows(min_row=2, max_col=2, values_only=True):
        assigned_person = row[1]
        if assigned_person in housemate_count:
            housemate_count[assigned_person] += 1

    # Update the 'Count' sheet with the counts
    for i, housemate in enumerate(housemates, start=2):
        ws3.cell(row=i, column=2, value=housemate_count[housemate["name"]])

    # Send email reminders for 1 week and 1 day before the collection date
    for row in ws1.iter_rows(min_row=2, max_col=6):
        collection_date = convert_excel_date(row[0].value)  # Convert to datetime.date
        assigned_person = row[1].value
        email = row[2].value
        reminder_1_week = row[4].value
        reminder_1_day = row[5].value

        # Ensure the collection_date is a datetime.date object before performing operations
        if isinstance(collection_date, str):
            collection_date = datetime.strptime(collection_date, "%d/%m/%Y").date()

        # Format the collection_date for the email (dd/mm/yyyy)
        formatted_date = format_date(collection_date)

        # Prepare the collection list (could be all bins for the collection day)
        #collection_list = row[3].value  # This assumes the collection names are in the 4th column (Collection Names)
        collection_list = "Recycling collection, Rubbish collection, Food waste collection"
        
        # Ensure collection_list is a string and split it by commas if needed (assuming comma-separated values)
        if isinstance(collection_list, str):
            collection_list_items = collection_list.split(',')  # Split by comma or change delimiter if needed
        else:
            collection_list_items = [str(collection_list)]  # In case it's not a string, we convert it to string

        # Format the collection list as a bulleted list
        collection_list_formatted = "\n".join([f"• {item.strip()}" for item in collection_list_items])

        # Prepare email body template with formatted date and collection list
        email_body = (
            f"Hello {assigned_person},\n\n"
            f"This is a reminder that you are assigned to take out the bins for collection day on {formatted_date}.\n\n"
            f"The following bins will be collected:\n{collection_list_formatted}\n\n"
            f"You can double check which bins here: {bin_collection_link}.\n\n"
            "The bins are collected early morning on collection day so make sure you take out the bins the night before.\n\n"
            "Thanks,\nHouse Bot"
        )

        # Send 1 week reminder
        if reminder_1_week == 'Not Sent' and (collection_date - current_date).days == 7:
            send_email(email, "Bin Collection Reminder - 1 Week", email_body)
            row[4].value = 'Sent'

        # Send 1 day reminder
        if reminder_1_day == 'Not Sent' and (collection_date - current_date).days == 1:
            send_email(email, "Bin Collection Reminder - 1 Day", email_body)
            row[5].value = 'Sent'

    # Save the updated workbook
    wb.save(file_path)

except Exception as e:
    exception_type = type(e).__name__
    exception_detail = str(e)
    print(f"Error occurred: {exception_type} - {exception_detail}")
    send_email(os.getenv("GMAIL_ADDRESS"), "Error in Bin Collection Assignment Script", 
               f"An error occurred in the script:\n\nException Type: {exception_type}\nException Detail: {exception_detail}")